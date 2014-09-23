# -*- coding: utf-8 -*-

from __future__ import unicode_literals

import csv
import json
from itertools import izip_longest

from django.utils.encoding import python_2_unicode_compatible
from django.utils import timezone
from django.db import models
from django.utils.translation import ugettext_lazy as _
from django.core.urlresolvers import reverse
from django_extensions.db.fields import UUIDField
from model_utils import Choices
from annoying.functions import get_object_or_None
from openpyxl import load_workbook

from categories.models import Category
from documents.models import Document
from documents.forms.models import documentform_factory
from documents.utils import save_document_forms


class normal_dialect(csv.Dialect):
    delimiter = b';'
    quotechar = b'"'
    doublequote = False
    skipinitialspace = True
    lineterminator = b'\r\n'
    quoting = csv.QUOTE_NONE
    strict = True
csv.register_dialect('normal', normal_dialect)


def xls_to_django(value):
    """Converts an excel value into a format we can use.

    Excel stores all numeric values as floats. For exemple, if you put "1" in
    a cell, importing that cell will yield a value of "1.0", which can cause
    errors.

    We need to convert numbers into strings since this is the format expected
    by django forms

    I feel like this is an awful hack. But Excel is a gigantic hack, so it's
    the best I can do.

    """
    if value is None:
        value = ''
    elif hasattr(value, 'is_integer') and value.is_integer():
        value = '%s' % int(value)
    else:
        value = '%s' % value

    return value


@python_2_unicode_compatible
class ImportBatch(models.Model):
    STATUSES = Choices(
        ('new', _('New')),
        ('started', _('Started')),
        ('success', _('Success')),
        ('partial_success', _('Partial success')),
        ('error', _('Error')),
    )

    uid = UUIDField(primary_key=True)
    category = models.ForeignKey(
        Category,
        verbose_name=_('Category')
    )
    file = models.FileField(
        _('File'),
        upload_to='import_%Y%m%d'
    )
    status = models.CharField(
        _('Status'),
        max_length=50,
        choices=STATUSES,
        default=STATUSES.new
    )
    created_on = models.DateField(
        _('Created on'),
        default=timezone.now
    )

    class Meta:
        verbose_name = _('Import batch')
        verbose_name_plural = _('Import batches')
        ordering = ['-created_on']

    @property
    def imported_type(self):
        return self.category.category_template.metadata_model

    def __str__(self):
        return 'Import {} ({})'.format(self.uid, self.imported_type)

    def get_absolute_url(self):
        return reverse('import_status', args=[self.uid])

    def get_form_class(self):
        form_class = documentform_factory(self.imported_type.model_class())
        return form_class

    def get_form(self, data=None, **kwargs):
        return self.get_form_class()(data, **kwargs)

    def get_revisionform_class(self):
        obj_class = self.imported_type.model_class()
        obj = obj_class()
        form_class = documentform_factory(obj.get_revision_class())
        return form_class

    def get_revisionform(self, data=None, **kwargs):
        return self.get_revisionform_class()(data, **kwargs)

    def __iter__(self):
        """Loop over csv data."""
        if self.file.path.endswith('csv'):
            with open(self.file.path, 'rb') as f:
                csvfile = csv.DictReader(f, dialect='normal')
                for row in csvfile:
                    imp = Import(batch=self, data=row)
                    yield imp
        else:
            wb = load_workbook(filename=self.file.path, use_iterators=True)
            sheet = wb.active
            header_row = None
            for row in sheet.iter_rows():
                if header_row is None:
                    header_row = [c.value for c in list(row)]
                    continue
                values = [xls_to_django(c.value) for c in list(row)]
                row = dict(izip_longest(header_row, values))
                imp = Import(batch=self, data=row)
                yield imp

    def do_import(self):
        line = 1
        error_count = 0
        for imp in self:
            imp.do_import(line)
            imp.save()
            if imp.status == Import.STATUSES.error:
                error_count += 1
            line += 1

        if error_count == line - 1:
            self.status = self.STATUSES.error
        elif error_count > 0:
            self.status = self.STATUSES.partial_success
        else:
            self.status = self.STATUSES.success
        self.save()


class Import(models.Model):
    STATUSES = Choices(
        ('new', _('New')),
        ('success', _('Success')),
        ('error', _('Error')),
    )

    line = models.IntegerField(_('Line'))
    batch = models.ForeignKey(
        ImportBatch,
        verbose_name=_('Batch'),
    )
    document = models.ForeignKey(
        Document,
        null=True, blank=True
    )
    status = models.CharField(
        _('Status'),
        max_length=50,
        choices=STATUSES,
        default=STATUSES.new
    )
    errors = models.TextField(
        _('Errors'),
        null=True, blank=True,
    )

    def __init__(self, *args, **kwargs):
        self.data = kwargs.pop('data', None)
        super(Import, self).__init__(*args, **kwargs)

    def get_forms(self, metadata_instance=None, revision_instance=None):
        return (
            self.batch.get_form(self.data, instance=metadata_instance),
            self.batch.get_revisionform(self.data, instance=revision_instance)
        )

    def do_import(self, line):
        assert hasattr(self, 'data')

        self.line = line

        # Checking if the document already exists
        key = self.data.get('document_key', None)
        doc = get_object_or_None(Document, document_key=key)
        metadata = doc.metadata if doc else None

        # Checking if the revision already exists
        revision_num = self.data.get('revision', None)
        revision = metadata.get_revision(revision_num) if metadata and revision_num else None

        form, revision_form = self.get_forms(metadata, revision)
        if form.is_valid() and revision_form.is_valid():
            doc, metadata, revision = save_document_forms(
                form, revision_form, self.batch.category)
            self.document = doc
            self.status = self.STATUSES.success
        else:
            errors = dict(form.errors.items() + revision_form.errors.items())
            self.errors = json.dumps(errors)
            self.status = self.STATUSES.error
